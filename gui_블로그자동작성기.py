import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import threading
import time
import pyperclip
import os
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from geminiapi import GeminiAPI
import shutil
import glob


class ModernBlogWriterGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("네이버 블로그 자동 작성기 v2.0")
        self.root.geometry("800x700")
        self.root.configure(bg='#f0f0f0')
        
        # 변수들
        self.driver = None
        self.gemini_api = None
        self.is_running = False
        self.current_step = ""
        
        self.setup_styles()
        self.create_widgets()
        self.center_window()
    
    def setup_styles(self):
        """스타일 설정"""
        self.colors = {
            'primary': '#2E86AB',
            'secondary': '#A23B72',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'danger': '#F44336',
            'light': '#f8f9fa',
            'dark': '#343a40'
        }
        
        style = ttk.Style()
        style.theme_use('clam')
        
        # 버튼 스타일
        style.configure('Primary.TButton',
                       background=self.colors['primary'],
                       foreground='white',
                       font=('맑은 고딕', 10, 'bold'),
                       padding=(10, 5))
        
        style.configure('Success.TButton',
                       background=self.colors['success'],
                       foreground='white',
                       font=('맑은 고딕', 10, 'bold'),
                       padding=(10, 5))
        
        style.configure('Warning.TButton',
                       background=self.colors['warning'],
                       foreground='white',
                       font=('맑은 고딕', 10, 'bold'),
                       padding=(10, 5))
    
    def create_widgets(self):
        """위젯 생성"""
        # 메인 프레임
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 제목
        title_label = tk.Label(main_frame, 
                              text="🚀 네이버 블로그 자동 작성기",
                              font=('맑은 고딕', 20, 'bold'),
                              fg=self.colors['primary'],
                              bg='#f0f0f0')
        title_label.pack(pady=(0, 20))
        
        # 탭 생성
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)
        
        # 탭 1: 로그인 설정
        self.create_login_tab()
        
        # 탭 2: SEO 키워드 작성
        self.create_seo_tab()
        
        # 탭 3: 엑셀 기반 작성
        self.create_excel_tab()
        
        # 탭 4: 로그
        self.create_log_tab()
        
        # 하단 상태바
        self.create_status_bar(main_frame)
    
    def create_login_tab(self):
        """로그인 설정 탭"""
        login_frame = ttk.Frame(self.notebook)
        self.notebook.add(login_frame, text="🔐 로그인 설정")
        
        # 로그인 정보 프레임
        login_info_frame = tk.LabelFrame(login_frame, 
                                        text="네이버 로그인 정보",
                                        font=('맑은 고딕', 12, 'bold'),
                                        bg='white',
                                        relief='solid',
                                        bd=1)
        login_info_frame.pack(fill='x', padx=20, pady=20)
        
        # 아이디 입력
        tk.Label(login_info_frame, text="네이버 아이디:", 
                font=('맑은 고딕', 10), bg='white').grid(row=0, column=0, sticky='w', padx=10, pady=5)
        self.id_entry = tk.Entry(login_info_frame, font=('맑은 고딕', 10), width=30)
        self.id_entry.grid(row=0, column=1, padx=10, pady=5)
        self.id_entry.insert(0, "cocodinglab")
        
        # 비밀번호 입력
        tk.Label(login_info_frame, text="비밀번호:", 
                font=('맑은 고딕', 10), bg='white').grid(row=1, column=0, sticky='w', padx=10, pady=5)
        self.pw_entry = tk.Entry(login_info_frame, font=('맑은 고딕', 10), width=30, show="*")
        self.pw_entry.grid(row=1, column=1, padx=10, pady=5)
        self.pw_entry.insert(0, "zhzheld201*")
        
        # 로그인 테스트 버튼
        test_login_btn = ttk.Button(login_info_frame, 
                                   text="🔍 로그인 테스트",
                                   style='Primary.TButton',
                                   command=self.test_login)
        test_login_btn.grid(row=2, column=0, columnspan=2, pady=10)
        
        # Gemini API 설정
        api_frame = tk.LabelFrame(login_frame, 
                                 text="Gemini API 설정",
                                 font=('맑은 고딕', 12, 'bold'),
                                 bg='white',
                                 relief='solid',
                                 bd=1)
        api_frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(api_frame, text="API 키:", 
                font=('맑은 고딕', 10), bg='white').grid(row=0, column=0, sticky='w', padx=10, pady=5)
        self.api_entry = tk.Entry(api_frame, font=('맑은 고딕', 10), width=50)
        self.api_entry.grid(row=0, column=1, padx=10, pady=5)
        self.api_entry.insert(0, "AIzaSyDj0ejAhXNSydhVHdcLDUFuirq4Xhy2B0I")
        
        test_api_btn = ttk.Button(api_frame, 
                                 text="🤖 API 테스트",
                                 style='Success.TButton',
                                 command=self.test_api)
        test_api_btn.grid(row=1, column=0, columnspan=2, pady=10)
    
    def create_seo_tab(self):
        """SEO 키워드 기반 작성 탭"""
        seo_frame = ttk.Frame(self.notebook)
        self.notebook.add(seo_frame, text="📝 SEO 키워드 작성")
        
        # 키워드 입력 프레임
        keyword_frame = tk.LabelFrame(seo_frame, 
                                     text="SEO 키워드 입력",
                                     font=('맑은 고딕', 12, 'bold'),
                                     bg='white',
                                     relief='solid',
                                     bd=1)
        keyword_frame.pack(fill='x', padx=20, pady=20)
        
        tk.Label(keyword_frame, text="키워드 (쉼표로 구분):", 
                font=('맑은 고딕', 10), bg='white').pack(anchor='w', padx=10, pady=5)
        
        self.keywords_entry = tk.Text(keyword_frame, height=3, font=('맑은 고딕', 10))
        self.keywords_entry.pack(fill='x', padx=10, pady=5)
        self.keywords_entry.insert('1.0', "코딩교육, 프로그래밍, 아이코딩, 진학, 특기자전형")
        
        # 글 생성 설정
        settings_frame = tk.LabelFrame(seo_frame, 
                                      text="글 생성 설정",
                                      font=('맑은 고딕', 12, 'bold'),
                                      bg='white',
                                      relief='solid',
                                      bd=1)
        settings_frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(settings_frame, text="생성할 글 개수:", 
                font=('맑은 고딕', 10), bg='white').grid(row=0, column=0, sticky='w', padx=10, pady=5)
        self.post_count_var = tk.StringVar(value="3")
        post_count_spin = tk.Spinbox(settings_frame, from_=1, to=10, 
                                    textvariable=self.post_count_var,
                                    font=('맑은 고딕', 10), width=10)
        post_count_spin.grid(row=0, column=1, padx=10, pady=5)
        
        # 자동 생성 및 포스팅 버튼
        btn_frame = tk.Frame(seo_frame, bg='#f0f0f0')
        btn_frame.pack(fill='x', padx=20, pady=20)
        
        generate_btn = ttk.Button(btn_frame, 
                                 text="✨ 글 자동 생성",
                                 style='Primary.TButton',
                                 command=self.generate_seo_posts)
        generate_btn.pack(side='left', padx=(0, 10))
        
        post_btn = ttk.Button(btn_frame, 
                             text="📤 생성 후 바로 포스팅",
                             style='Success.TButton',
                             command=self.generate_and_post)
        post_btn.pack(side='left')
    
    def create_excel_tab(self):
        """엑셀 기반 작성 탭"""
        excel_frame = ttk.Frame(self.notebook)
        self.notebook.add(excel_frame, text="📊 엑셀 기반 작성")
        
        # 파일 선택 프레임
        file_frame = tk.LabelFrame(excel_frame, 
                                  text="엑셀 파일 선택",
                                  font=('맑은 고딕', 12, 'bold'),
                                  bg='white',
                                  relief='solid',
                                  bd=1)
        file_frame.pack(fill='x', padx=20, pady=20)
        
        tk.Label(file_frame, text="선택된 파일:", 
                font=('맑은 고딕', 10), bg='white').grid(row=0, column=0, sticky='w', padx=10, pady=5)
        self.file_path_var = tk.StringVar(value="posting.xlsx")
        file_label = tk.Label(file_frame, textvariable=self.file_path_var,
                             font=('맑은 고딕', 10), bg='white', fg=self.colors['primary'])
        file_label.grid(row=0, column=1, sticky='w', padx=10, pady=5)
        
        browse_btn = ttk.Button(file_frame, 
                               text="📂 파일 선택",
                               style='Primary.TButton',
                               command=self.browse_file)
        browse_btn.grid(row=1, column=0, columnspan=2, pady=10)
        
        # 엑셀 처리 버튼
        excel_btn_frame = tk.Frame(excel_frame, bg='#f0f0f0')
        excel_btn_frame.pack(fill='x', padx=20, pady=20)
        
        process_btn = ttk.Button(excel_btn_frame, 
                                text="🚀 엑셀 데이터로 자동 포스팅",
                                style='Success.TButton',
                                command=self.process_excel)
        process_btn.pack()
    
    def create_log_tab(self):
        """로그 탭"""
        log_frame = ttk.Frame(self.notebook)
        self.notebook.add(log_frame, text="📋 실행 로그")
        
        # 로그 텍스트 영역
        self.log_text = scrolledtext.ScrolledText(log_frame, 
                                                 font=('Consolas', 10),
                                                 bg='#2d3748',
                                                 fg='#e2e8f0',
                                                 insertbackground='white')
        self.log_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 로그 제어 버튼
        log_btn_frame = tk.Frame(log_frame, bg='#f0f0f0')
        log_btn_frame.pack(fill='x', padx=10, pady=5)
        
        clear_btn = ttk.Button(log_btn_frame, 
                              text="🗑️ 로그 지우기",
                              style='Warning.TButton',
                              command=self.clear_log)
        clear_btn.pack(side='right')
    
    def create_status_bar(self, parent):
        """상태바 생성"""
        status_frame = tk.Frame(parent, bg=self.colors['dark'], height=30)
        status_frame.pack(fill='x', side='bottom')
        status_frame.pack_propagate(False)
        
        self.status_var = tk.StringVar(value="준비됨")
        status_label = tk.Label(status_frame, 
                               textvariable=self.status_var,
                               bg=self.colors['dark'],
                               fg='white',
                               font=('맑은 고딕', 9))
        status_label.pack(side='left', padx=10, pady=5)
        
        # 진행률 바
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate')
        self.progress.pack(side='right', padx=10, pady=5, fill='x', expand=True)
    
    def center_window(self):
        """창을 화면 중앙에 배치"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def log(self, message, level='INFO'):
        """로그 출력"""
        timestamp = time.strftime("%H:%M:%S")
        log_colors = {
            'INFO': '#4CAF50',
            'WARNING': '#FF9800', 
            'ERROR': '#F44336',
            'SUCCESS': '#2196F3'
        }
        
        self.log_text.insert('end', f"[{timestamp}] [{level}] {message}\n")
        self.log_text.see('end')
        self.root.update()
    
    def clear_log(self):
        """로그 지우기"""
        self.log_text.delete(1.0, 'end')
    
    def update_status(self, status):
        """상태 업데이트"""
        self.status_var.set(status)
        self.root.update()
    
    def setup_driver(self):
        """Chrome 드라이버 설정"""
        try:
            self.log("Chrome 드라이버 설정 중...")
            self.update_status("드라이버 초기화 중...")
            
            # 기존 드라이버 파일 정리
            temp_paths = [
                os.path.expanduser("~/.wdm"),
                os.path.join(os.getcwd(), "chromedriver*"),
            ]
            
            for pattern in temp_paths:
                try:
                    for path in glob.glob(pattern):
                        if os.path.exists(path):
                            if os.path.isfile(path):
                                os.remove(path)
                            elif os.path.isdir(path):
                                shutil.rmtree(path)
                except:
                    pass
            
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            self.log("새로운 Chrome 드라이버 다운로드 중...")
            driver_path = ChromeDriverManager().install()
            
            service = Service(driver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            self.log("Chrome 드라이버 설정 완료!", 'SUCCESS')
            return True
            
        except Exception as e:
            self.log(f"드라이버 설정 중 오류: {e}", 'ERROR')
            return False
    
    def test_login(self):
        """로그인 테스트"""
        def test_thread():
            try:
                self.progress.start()
                self.update_status("로그인 테스트 중...")
                
                if not self.setup_driver():
                    return
                
                naver_id = self.id_entry.get().strip()
                naver_pw = self.pw_entry.get().strip()
                
                if not naver_id or not naver_pw:
                    messagebox.showerror("오류", "아이디와 비밀번호를 입력해주세요.")
                    return
                
                self.log("네이버 로그인 페이지로 이동 중...")
                self.driver.get("https://nid.naver.com/nidlogin.login")
                time.sleep(3)
                
                # 아이디 입력
                id_input = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "id"))
                )
                pyperclip.copy(naver_id)
                id_input.click()
                time.sleep(0.5)
                id_input.clear()
                actions = ActionChains(self.driver)
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                
                # 비밀번호 입력
                pw_input = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "pw"))
                )
                pyperclip.copy(naver_pw)
                pw_input.click()
                time.sleep(0.5)
                pw_input.clear()
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                
                # 로그인 버튼 클릭
                login_btn = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "log.login"))
                )
                login_btn.click()
                time.sleep(5)
                
                current_url = self.driver.current_url
                if ("naver.com" in current_url and "login" not in current_url):
                    self.log("✅ 로그인 성공!", 'SUCCESS')
                    messagebox.showinfo("성공", "로그인이 성공했습니다!")
                else:
                    self.log("❌ 로그인 실패", 'ERROR')
                    messagebox.showerror("실패", "로그인에 실패했습니다.")
                
            except Exception as e:
                self.log(f"로그인 테스트 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                if self.driver:
                    self.driver.quit()
                    self.driver = None
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=test_thread, daemon=True).start()
    
    def test_api(self):
        """API 테스트"""
        def api_thread():
            try:
                self.progress.start()
                self.update_status("API 테스트 중...")
                
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.log("Gemini API 초기화 중...")
                self.gemini_api = GeminiAPI(api_key)
                
                self.log("API 테스트 요청 중...")
                test_prompt = "안녕하세요. API 테스트입니다. 간단히 인사해주세요."
                response = self.gemini_api.generate_text(test_prompt)
                
                if response:
                    self.log("✅ API 테스트 성공!", 'SUCCESS')
                    messagebox.showinfo("성공", f"API 테스트 성공!\n응답: {response[:100]}...")
                else:
                    self.log("❌ API 테스트 실패", 'ERROR')
                    messagebox.showerror("실패", "API 테스트에 실패했습니다.")
                
            except Exception as e:
                self.log(f"API 테스트 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"API 오류: {e}")
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=api_thread, daemon=True).start()
    
    def generate_seo_posts(self):
        """SEO 키워드 기반 글 생성"""
        def generate_thread():
            try:
                self.progress.start()
                self.update_status("글 생성 중...")
                
                # API 초기화
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.gemini_api = GeminiAPI(api_key)
                
                # 키워드 처리
                keywords_text = self.keywords_entry.get('1.0', 'end').strip()
                keywords = [k.strip() for k in keywords_text.split(',') if k.strip()]
                
                if not keywords:
                    messagebox.showerror("오류", "키워드를 입력해주세요.")
                    return
                
                post_count = int(self.post_count_var.get())
                
                self.log(f"키워드: {', '.join(keywords)}")
                self.log(f"생성할 글 개수: {post_count}")
                
                # 엑셀 파일 생성
                wb = Workbook()
                ws = wb.active
                ws.title = "생성된 포스트"
                ws['A1'] = "제목"
                ws['B1'] = "본문"
                
                for i in range(post_count):
                    self.log(f"글 {i+1}/{post_count} 생성 중...")
                    
                    # 제목 생성
                    title_prompt = f"다음 키워드들을 활용한 SEO 최적화된 블로그 제목을 1개 생성해주세요: {', '.join(keywords)}. 제목만 출력하세요."
                    title = self.gemini_api.generate_text(title_prompt, temperature=0.8)
                    
                    if title:
                        title = title.strip().replace('"', '').replace("'", "")
                        
                        # 본문 생성
                        content_prompt = f"""
                        제목: "{title}"
                        키워드: {', '.join(keywords)}
                        
                        위 제목과 키워드를 바탕으로 SEO에 최적화된 블로그 글을 작성해주세요.
                        
                        구조:
                        1. 서론 (2-3문단): 독자의 관심을 끄는 도입부
                        2. 본론 (4-6문단): 키워드를 자연스럽게 포함한 핵심 내용
                        3. 결론 (1-2문단): 요약 및 행동 유도
                        
                        - 1500-2000자 분량
                        - 키워드를 자연스럽게 3-5회 포함
                        - 읽기 쉽고 유익한 내용
                        """
                        
                        content = self.gemini_api.generate_text(content_prompt, temperature=0.7, max_output_tokens=2048)
                        
                        if content:
                            ws[f'A{i+2}'] = title
                            ws[f'B{i+2}'] = content
                            self.log(f"✅ 글 {i+1} 생성 완료: {title[:30]}...")
                        else:
                            self.log(f"❌ 글 {i+1} 본문 생성 실패", 'WARNING')
                    else:
                        self.log(f"❌ 글 {i+1} 제목 생성 실패", 'WARNING')
                    
                    time.sleep(1)  # API 제한 고려
                
                # 파일 저장
                filename = f"generated_posts_{int(time.time())}.xlsx"
                wb.save(filename)
                wb.close()
                
                self.log(f"✅ 모든 글 생성 완료! 파일: {filename}", 'SUCCESS')
                messagebox.showinfo("완료", f"글 생성이 완료되었습니다!\n저장된 파일: {filename}")
                
            except Exception as e:
                self.log(f"글 생성 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=generate_thread, daemon=True).start()
    
    def generate_and_post(self):
        """글 생성 후 바로 포스팅"""
        def generate_post_thread():
            try:
                self.progress.start()
                self.update_status("글 생성 및 포스팅 중...")
                
                # 먼저 글 생성
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.gemini_api = GeminiAPI(api_key)
                
                keywords_text = self.keywords_entry.get('1.0', 'end').strip()
                keywords = [k.strip() for k in keywords_text.split(',') if k.strip()]
                
                if not keywords:
                    messagebox.showerror("오류", "키워드를 입력해주세요.")
                    return
                
                # 드라이버 설정
                if not self.setup_driver():
                    return
                
                # 로그인
                if not self.login_to_naver():
                    return
                
                post_count = int(self.post_count_var.get())
                
                for i in range(post_count):
                    self.log(f"글 {i+1}/{post_count} 생성 및 포스팅 중...")
                    
                    # 제목 생성
                    title_prompt = f"다음 키워드들을 활용한 SEO 최적화된 블로그 제목을 1개 생성해주세요: {', '.join(keywords)}. 제목만 출력하세요."
                    title = self.gemini_api.generate_text(title_prompt, temperature=0.8)
                    
                    if title:
                        title = title.strip().replace('"', '').replace("'", "")
                        
                        # 본문 생성
                        content_prompt = f"""
                        제목: "{title}"
                        키워드: {', '.join(keywords)}
                        
                        위 제목과 키워드를 바탕으로 블로그 글을 작성해주세요. 1500-2000자 분량으로 SEO에 최적화된 내용을 작성하세요.
                        """
                        
                        content = self.gemini_api.generate_text(content_prompt, temperature=0.7, max_output_tokens=2048)
                        
                        if content:
                            # 바로 포스팅
                            if self.write_blog_post(title, content):
                                self.log(f"✅ 글 {i+1} 포스팅 완료: {title[:30]}...")
                            else:
                                self.log(f"❌ 글 {i+1} 포스팅 실패", 'WARNING')
                        else:
                            self.log(f"❌ 글 {i+1} 생성 실패", 'WARNING')
                    
                    # 다음 글을 위한 대기
                    if i < post_count - 1:
                        self.log("다음 글 준비를 위해 3초 대기...")
                        time.sleep(3)
                
                self.log(f"✅ 모든 글 포스팅 완료!", 'SUCCESS')
                messagebox.showinfo("완료", "모든 글이 성공적으로 포스팅되었습니다!")
                
            except Exception as e:
                self.log(f"오류 발생: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                if self.driver:
                    self.driver.quit()
                    self.driver = None
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=generate_post_thread, daemon=True).start()
    
    def browse_file(self):
        """파일 선택"""
        filename = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
    
    def process_excel(self):
        """엑셀 파일 처리"""
        def excel_thread():
            try:
                self.progress.start()
                self.update_status("엑셀 데이터 처리 중...")
                
                excel_file = self.file_path_var.get()
                if not os.path.exists(excel_file):
                    messagebox.showerror("오류", "엑셀 파일을 찾을 수 없습니다.")
                    return
                
                # 드라이버 설정
                if not self.setup_driver():
                    return
                
                # 로그인
                if not self.login_to_naver():
                    return
                
                # 엑셀 데이터 로드
                posts_data = self.load_excel_data(excel_file)
                if not posts_data:
                    return
                
                success_count = 0
                for i, post in enumerate(posts_data, 1):
                    self.log(f"[{i}/{len(posts_data)}] {post['row']}행 처리 중...")
                    
                    if self.write_blog_post(post['title'], post['content']):
                        success_count += 1
                        self.log(f"✅ {post['row']}행 포스팅 성공!")
                    else:
                        self.log(f"❌ {post['row']}행 포스팅 실패", 'WARNING')
                    
                    if i < len(posts_data):
                        time.sleep(3)
                
                self.log(f"✅ 처리 완료! 성공: {success_count}/{len(posts_data)}", 'SUCCESS')
                messagebox.showinfo("완료", f"엑셀 데이터 처리 완료!\n성공: {success_count}/{len(posts_data)}")
                
            except Exception as e:
                self.log(f"엑셀 처리 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                if self.driver:
                    self.driver.quit()
                    self.driver = None
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=excel_thread, daemon=True).start()
    
    def load_excel_data(self, filename):
        """엑셀 데이터 로드"""
        try:
            self.log(f"엑셀 파일 로드: {filename}")
            workbook = load_workbook(filename)
            worksheet = workbook.active
            
            posts_data = []
            max_row = worksheet.max_row
            
            for row_num in range(2, max_row + 1):
                title_cell = worksheet[f'A{row_num}']
                content_cell = worksheet[f'B{row_num}']
                
                if not title_cell.value or title_cell.value.strip() == "":
                    continue
                
                title = title_cell.value.strip()
                content = content_cell.value.strip() if content_cell.value else ""
                
                posts_data.append({
                    'row': row_num,
                    'title': title,
                    'content': content
                })
            
            workbook.close()
            self.log(f"총 {len(posts_data)}개의 게시글 로드 완료")
            return posts_data
            
        except Exception as e:
            self.log(f"엑셀 파일 로드 중 오류: {e}", 'ERROR')
            return []
    
    def login_to_naver(self):
        """네이버 로그인"""
        try:
            self.log("네이버 로그인 시작...")
            
            naver_id = self.id_entry.get().strip()
            naver_pw = self.pw_entry.get().strip()
            
            if not naver_id or not naver_pw:
                messagebox.showerror("오류", "아이디와 비밀번호를 입력해주세요.")
                return False
            
            self.driver.get("https://nid.naver.com/nidlogin.login")
            time.sleep(3)
            
            # 아이디 입력
            id_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "id"))
            )
            pyperclip.copy(naver_id)
            id_input.click()
            time.sleep(0.5)
            id_input.clear()
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 비밀번호 입력
            pw_input = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.ID, "pw"))
            )
            pyperclip.copy(naver_pw)
            pw_input.click()
            time.sleep(0.5)
            pw_input.clear()
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 로그인 버튼 클릭
            login_btn = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.ID, "log.login"))
            )
            login_btn.click()
            time.sleep(5)
            
            current_url = self.driver.current_url
            if ("naver.com" in current_url and "login" not in current_url):
                self.log("✅ 로그인 성공!", 'SUCCESS')
                return True
            else:
                self.log("❌ 로그인 실패", 'ERROR')
                return False
                
        except Exception as e:
            self.log(f"로그인 중 오류: {e}", 'ERROR')
            return False
    
    def write_blog_post(self, title, content):
        """블로그 포스트 작성"""
        try:
            self.log(f"블로그 포스트 작성: {title[:30]}...")
            
            # 글쓰기 페이지로 이동
            self.driver.get("https://blog.naver.com/GoBlogWrite.naver")
            time.sleep(3)
            
            # iframe 전환
            main_frame = WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#mainFrame"))
            )
            self.driver.switch_to.frame(main_frame)
            time.sleep(2)
            
            # 팝업 처리
            self.handle_popups()
            
            # 제목 입력
            title_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-documentTitle"))
            )
            title_element.click()
            time.sleep(1)
            
            pyperclip.copy(title)
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.3)
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 본문 입력
            text_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-text"))
            )
            text_element.click()
            time.sleep(1)
            
            if content:
                pyperclip.copy(content)
                actions = ActionChains(self.driver)
                actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                time.sleep(0.3)
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(2)
            
            # 저장 버튼 클릭
            save_selectors = [
                ".save_btn__bzc5B",
                "#save_btn_bzc58", 
                ".btn_save",
                "[class*='save']"
            ]
            
            for selector in save_selectors:
                try:
                    save_btn = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    save_btn.click()
                    self.log(f"저장 버튼 클릭 완료!")
                    break
                except:
                    continue
            
            time.sleep(3)
            self.driver.switch_to.default_content()
            
            return True
            
        except Exception as e:
            self.log(f"글 작성 중 오류: {e}", 'ERROR')
            try:
                self.driver.switch_to.default_content()
            except:
                pass
            return False
    
    def handle_popups(self):
        """팝업 처리"""
        try:
            cancel_btn = self.driver.find_element(By.CSS_SELECTOR, ".se-popup-button-cancel")
            if cancel_btn.is_displayed():
                cancel_btn.click()
                time.sleep(1)
        except:
            pass
        
        try:
            help_close_btn = self.driver.find_element(By.CSS_SELECTOR, ".se-help-panel-close-button")
            if help_close_btn.is_displayed():
                help_close_btn.click()
                time.sleep(1)
        except:
            pass
    
    def run(self):
        """GUI 실행"""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()
    
    def on_closing(self):
        """프로그램 종료 시 처리"""
        if self.driver:
            self.driver.quit()
        self.root.destroy()


if __name__ == "__main__":
    app = ModernBlogWriterGUI()
    app.run()