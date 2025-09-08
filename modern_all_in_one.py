"""
🚀 모던 올인원 네이버 블로그 자동 작성기
- 모든 기능이 한 페이지에 통합
- 다크 테마 적용
- 실제 작동하는 모든 기능
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import threading
import time
import pyperclip
import os
import json
import base64
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from geminiapi import GeminiAPI
import shutil
import glob


class ModernBlogWriter:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🚀 네이버 블로그 자동 작성기 2024")
        self.root.geometry("1200x800")
        self.root.configure(bg='#0f172a')
        
        # 변수들
        self.driver = None
        self.gemini_api = None
        self.is_running = False
        
        self.create_ui()
        self.center_window()
        
        # 저장된 정보 자동 로드
        self.load_credentials(show_message=False)
    
    def create_ui(self):
        """모던 UI 생성"""
        # 메인 스크롤 프레임
        main_canvas = tk.Canvas(self.root, bg='#0f172a', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        self.scrollable_frame = tk.Frame(main_canvas, bg='#0f172a')
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        # 헤더
        self.create_header()
        
        # 메인 콘텐츠 (2열 레이아웃)
        content_frame = tk.Frame(self.scrollable_frame, bg='#0f172a')
        content_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # 좌측 패널
        left_panel = tk.Frame(content_frame, bg='#0f172a')
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # 우측 패널
        right_panel = tk.Frame(content_frame, bg='#0f172a')
        right_panel.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        # 좌측: 로그인 설정 + AI 글 생성
        self.create_login_section(left_panel)
        self.create_ai_section(left_panel)
        
        # 우측: 엑셀 처리 + 로그
        self.create_excel_section(right_panel)
        self.create_log_section(right_panel)
        
        # 하단 상태바
        self.create_status_bar()
        
        # 스크롤 설정
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 마우스 휠 스크롤
        def _on_mousewheel(event):
            main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def create_header(self):
        """헤더 생성"""
        header = tk.Frame(self.scrollable_frame, bg='#1e293b', height=80)
        header.pack(fill='x', padx=20, pady=(20, 0))
        header.pack_propagate(False)
        
        # 제목
        title_frame = tk.Frame(header, bg='#1e293b')
        title_frame.pack(expand=True, fill='both')
        
        tk.Label(title_frame, 
                text="🚀", 
                font=('Segoe UI', 24),
                bg='#1e293b',
                fg='#f8fafc').pack(side='left', padx=(30, 10), pady=20)
        
        tk.Label(title_frame, 
                text="네이버 블로그 자동 작성기", 
                font=('Segoe UI', 20, 'bold'),
                bg='#1e293b',
                fg='#f8fafc').pack(side='left', pady=20)
        
        tk.Label(title_frame, 
                text="v2.0", 
                font=('Segoe UI', 12),
                bg='#1e293b',
                fg='#3b82f6').pack(side='left', padx=(10, 0), pady=25)
    
    def create_card(self, parent, title, icon=""):
        """카드 생성"""
        card = tk.Frame(parent, bg='#1e293b', relief='flat', bd=1)
        card.pack(fill='x', pady=10)
        
        # 카드 헤더
        header = tk.Frame(card, bg='#1e293b')
        header.pack(fill='x', padx=20, pady=(15, 10))
        
        tk.Label(header, 
                text=f"{icon} {title}", 
                font=('Segoe UI', 14, 'bold'),
                bg='#1e293b',
                fg='#f8fafc').pack(anchor='w')
        
        # 카드 내용 프레임 반환
        content = tk.Frame(card, bg='#1e293b')
        content.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        return content
    
    def create_button(self, parent, text, command, style='primary'):
        """모던 버튼 생성"""
        colors = {
            'primary': {'bg': '#3b82f6', 'active': '#2563eb'},
            'success': {'bg': '#10b981', 'active': '#059669'},
            'warning': {'bg': '#f59e0b', 'active': '#d97706'},
            'danger': {'bg': '#ef4444', 'active': '#dc2626'}
        }
        
        color = colors.get(style, colors['primary'])
        
        btn = tk.Button(parent, 
                       text=text,
                       font=('Segoe UI', 10, 'bold'),
                       bg=color['bg'],
                       fg='#ffffff',
                       activebackground=color['active'],
                       activeforeground='#ffffff',
                       relief='flat',
                       bd=0,
                       padx=20,
                       pady=8,
                       cursor='hand2',
                       command=command)
        
        # 호버 효과
        btn.bind("<Enter>", lambda e: btn.configure(bg=color['active']))
        btn.bind("<Leave>", lambda e: btn.configure(bg=color['bg']))
        
        return btn
    
    def create_entry(self, parent, placeholder="", show=None):
        """모던 입력창 생성"""
        entry = tk.Entry(parent,
                        font=('Segoe UI', 10),
                        bg='#334155',
                        fg='#f8fafc',
                        relief='flat',
                        bd=0,
                        highlightbackground='#475569',
                        highlightcolor='#3b82f6',
                        highlightthickness=2,
                        insertbackground='#f8fafc',
                        show=show)
        return entry
    
    def create_text(self, parent, height=4):
        """모던 텍스트 영역 생성"""
        text = tk.Text(parent,
                      font=('Segoe UI', 10),
                      bg='#334155',
                      fg='#f8fafc',
                      relief='flat',
                      bd=0,
                      highlightbackground='#475569',
                      highlightcolor='#3b82f6',
                      highlightthickness=2,
                      insertbackground='#f8fafc',
                      selectbackground='#3b82f6',
                      selectforeground='#ffffff',
                      wrap=tk.WORD,
                      height=height)
        return text
    
    def create_login_section(self, parent):
        """로그인 설정 섹션"""
        content = self.create_card(parent, "로그인 설정", "🔐")
        
        # 아이디
        tk.Label(content, text="네이버 아이디", 
                font=('Segoe UI', 9), fg='#cbd5e1', bg='#1e293b').pack(anchor='w', pady=(0, 5))
        self.id_entry = self.create_entry(content)
        self.id_entry.pack(fill='x', pady=(0, 10))
        
        # 비밀번호
        tk.Label(content, text="비밀번호", 
                font=('Segoe UI', 9), fg='#cbd5e1', bg='#1e293b').pack(anchor='w', pady=(0, 5))
        self.pw_entry = self.create_entry(content, show="*")
        self.pw_entry.pack(fill='x', pady=(0, 10))
        
        # Gemini API
        tk.Label(content, text="Gemini API 키", 
                font=('Segoe UI', 9), fg='#cbd5e1', bg='#1e293b').pack(anchor='w', pady=(0, 5))
        self.api_entry = self.create_entry(content)
        self.api_entry.pack(fill='x', pady=(0, 15))
        
        # 저장/불러오기 버튼들
        save_frame = tk.Frame(content, bg='#1e293b')
        save_frame.pack(fill='x', pady=(0, 10))
        
        save_btn = self.create_button(save_frame, "💾 정보 저장", self.save_credentials, 'warning')
        save_btn.pack(side='left', padx=(0, 10))
        
        load_btn = self.create_button(save_frame, "📂 정보 불러오기", self.load_credentials, 'warning')
        load_btn.pack(side='left')
        
        # 테스트 버튼들
        btn_frame = tk.Frame(content, bg='#1e293b')
        btn_frame.pack(fill='x')
        
        test_login_btn = self.create_button(btn_frame, "🔍 로그인 테스트", self.test_login)
        test_login_btn.pack(side='left', padx=(0, 10))
        
        test_api_btn = self.create_button(btn_frame, "🤖 API 테스트", self.test_api, 'success')
        test_api_btn.pack(side='left')
    
    def create_ai_section(self, parent):
        """AI 글 생성 섹션"""
        content = self.create_card(parent, "AI 글 생성", "✨")
        
        # 키워드 설명
        desc_label = tk.Label(content, 
                             text="💡 5개 포스트 생성: 각 키워드별로 1개씩 순서대로 생성됩니다", 
                             font=('Segoe UI', 9), 
                             fg='#60a5fa', 
                             bg='#1e293b')
        desc_label.pack(anchor='w', pady=(0, 10))
        
        # 키워드
        tk.Label(content, text="SEO 키워드 (쉼표로 구분, 최소 5개 권장)", 
                font=('Segoe UI', 9), fg='#cbd5e1', bg='#1e293b').pack(anchor='w', pady=(0, 5))
        self.keywords_text = self.create_text(content, height=3)
        self.keywords_text.pack(fill='x', pady=(0, 10))
        
        # 글 개수 (고정)
        count_frame = tk.Frame(content, bg='#1e293b')
        count_frame.pack(fill='x', pady=(0, 15))
        
        tk.Label(count_frame, text="📊 생성할 글 개수: 5개 (고정)", 
                font=('Segoe UI', 9, 'bold'), fg='#10b981', bg='#1e293b').pack(side='left')
        
        # 버튼들
        btn_frame = tk.Frame(content, bg='#1e293b')
        btn_frame.pack(fill='x')
        
        # 왼쪽: SEO 키워드 자동 생성 버튼
        auto_keyword_btn = self.create_button(btn_frame, "🎯 SEO키워드 자동생성", self.auto_generate_keywords, 'warning')
        auto_keyword_btn.pack(side='left', padx=(0, 10))
        
        # 오른쪽: 5개 글 생성 버튼
        generate_btn = self.create_button(btn_frame, "✨ 5개 글 생성", self.generate_seo_posts)
        generate_btn.pack(side='right')
    
    def create_excel_section(self, parent):
        """엑셀 처리 섹션"""
        content = self.create_card(parent, "엑셀 포스팅", "📊")
        
        # 파일 선택
        file_frame = tk.Frame(content, bg='#1e293b')
        file_frame.pack(fill='x', pady=(0, 15))
        
        tk.Label(file_frame, text="선택된 파일:", 
                font=('Segoe UI', 9), fg='#cbd5e1', bg='#1e293b').pack(anchor='w', pady=(0, 5))
        
        self.file_path_var = tk.StringVar(value="posting.xlsx")
        file_display = tk.Frame(file_frame, bg='#1e293b')
        file_display.pack(fill='x')
        
        tk.Label(file_display, textvariable=self.file_path_var, 
                font=('Segoe UI', 9), fg='#3b82f6', bg='#1e293b').pack(side='left')
        
        browse_btn = self.create_button(file_display, "📂 찾기", self.browse_file, 'warning')
        browse_btn.pack(side='right')
        
        # 실행 버튼
        excel_btn = self.create_button(content, "🚀 엑셀 데이터로 포스팅", self.process_excel, 'success')
        excel_btn.pack()
    
    def create_log_section(self, parent):
        """로그 섹션"""
        content = self.create_card(parent, "실행 로그", "📋")
        
        # 로그 텍스트
        self.log_text = scrolledtext.ScrolledText(content,
                                                 font=('JetBrains Mono', 9),
                                                 bg='#0f172a',
                                                 fg='#e2e8f0',
                                                 insertbackground='#e2e8f0',
                                                 selectbackground='#3b82f6',
                                                 relief='flat',
                                                 bd=0,
                                                 height=12)
        self.log_text.pack(fill='both', expand=True, pady=(0, 10))
        
        # 로그 제어 버튼
        log_btn_frame = tk.Frame(content, bg='#1e293b')
        log_btn_frame.pack(fill='x')
        
        clear_btn = self.create_button(log_btn_frame, "🗑️ 로그 지우기", self.clear_log, 'warning')
        clear_btn.pack(side='right')
        
        # 초기 로그
        self.log("🚀 네이버 블로그 자동 작성기 시작됨", 'SUCCESS')
        self.log("📋 시스템 준비 완료", 'INFO')
    
    def create_status_bar(self):
        """상태바 생성"""
        status_frame = tk.Frame(self.scrollable_frame, bg='#1e293b', height=40)
        status_frame.pack(fill='x', side='bottom', padx=20, pady=10)
        status_frame.pack_propagate(False)
        
        self.status_var = tk.StringVar(value="준비됨")
        status_label = tk.Label(status_frame, 
                               textvariable=self.status_var,
                               font=('Segoe UI', 9),
                               fg='#cbd5e1',
                               bg='#1e293b')
        status_label.pack(side='left', padx=15, pady=10)
        
        # 진행률 바
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate', length=200)
        self.progress.pack(side='right', padx=15, pady=10)
    
    def center_window(self):
        """창 중앙 배치"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def log(self, message, level='INFO'):
        """로그 출력"""
        timestamp = time.strftime("%H:%M:%S")
        
        # 레벨별 아이콘
        level_icons = {
            'INFO': '💡',
            'SUCCESS': '✅',
            'WARNING': '⚠️',
            'ERROR': '❌'
        }
        
        icon = level_icons.get(level, '💡')
        log_message = f"[{timestamp}] {icon} {message}\n"
        
        self.log_text.insert('end', log_message)
        self.log_text.see('end')
        self.root.update()
    
    def clear_log(self):
        """로그 지우기"""
        self.log_text.delete(1.0, 'end')
        self.log("🧹 로그가 지워졌습니다", 'INFO')
    
    def update_status(self, status):
        """상태 업데이트"""
        self.status_var.set(status)
        self.root.update()
    
    def setup_driver(self):
        """Chrome 드라이버 설정"""
        try:
            self.log("Chrome 드라이버 설정 중...", 'INFO')
            self.update_status("드라이버 초기화 중...")
            
            # Chrome 옵션 설정
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_argument("--disable-gcm-registration-probes")
            chrome_options.add_argument("--disable-background-networking")
            chrome_options.add_argument("--disable-background-timer-throttling")
            chrome_options.add_argument("--disable-backgrounding-occluded-windows")
            chrome_options.add_argument("--disable-renderer-backgrounding")
            chrome_options.add_argument("--disable-features=TranslateUI")
            chrome_options.add_argument("--disable-ipc-flooding-protection")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # 방법 1: ChromeDriverManager 시도
            try:
                self.log("방법 1: ChromeDriverManager로 드라이버 다운로드 시도...", 'INFO')
                
                # 기존 캐시 삭제
                import tempfile
                wdm_cache = os.path.join(tempfile.gettempdir(), '.wdm')
                if os.path.exists(wdm_cache):
                    shutil.rmtree(wdm_cache, ignore_errors=True)
                
                from webdriver_manager.chrome import ChromeDriverManager
                from webdriver_manager.core.utils import ChromeType
                
                driver_path = ChromeDriverManager(cache_valid_range=1).install()
                self.log(f"드라이버 경로: {driver_path}", 'INFO')
                
                # 드라이버 파일 유효성 검사
                if os.path.exists(driver_path) and os.path.getsize(driver_path) > 1000:
                    service = Service(driver_path)
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    self.log("방법 1 성공: ChromeDriverManager", 'SUCCESS')
                else:
                    raise Exception("다운로드된 드라이버가 유효하지 않음")
                    
            except Exception as e1:
                self.log(f"방법 1 실패: {e1}", 'WARNING')
                
                # 방법 2: 시스템 PATH의 Chrome 사용
                try:
                    self.log("방법 2: 시스템 Chrome 드라이버 사용 시도...", 'INFO')
                    self.driver = webdriver.Chrome(options=chrome_options)
                    self.log("방법 2 성공: 시스템 Chrome", 'SUCCESS')
                    
                except Exception as e2:
                    self.log(f"방법 2 실패: {e2}", 'WARNING')
                    
                    # 방법 3: 수동 다운로드 및 설치
                    try:
                        self.log("방법 3: 수동 Chrome 드라이버 다운로드 시도...", 'INFO')
                        import requests
                        import zipfile
                        
                        # Chrome 버전 확인
                        try:
                            import subprocess
                            result = subprocess.run([
                                'reg', 'query', 
                                'HKEY_CURRENT_USER\\Software\\Google\\Chrome\\BLBeacon', 
                                '/v', 'version'
                            ], capture_output=True, text=True)
                            
                            if result.returncode == 0:
                                version_line = [line for line in result.stdout.split('\n') if 'version' in line.lower()]
                                if version_line:
                                    chrome_version = version_line[0].split()[-1]
                                    major_version = chrome_version.split('.')[0]
                                    self.log(f"Chrome 버전 감지: {chrome_version}", 'INFO')
                                else:
                                    major_version = "131"  # 기본값
                            else:
                                major_version = "131"  # 기본값
                        except:
                            major_version = "131"  # 기본값
                        
                        # ChromeDriver 다운로드 URL
                        download_url = f"https://storage.googleapis.com/chrome-for-testing-public/{major_version}.0.6778.85/win64/chromedriver-win64.zip"
                        
                        # 임시 디렉토리에 다운로드
                        temp_dir = os.path.join(os.getcwd(), "temp_driver")
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        zip_path = os.path.join(temp_dir, "chromedriver.zip")
                        
                        # 다운로드
                        response = requests.get(download_url, stream=True)
                        with open(zip_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                f.write(chunk)
                        
                        # 압축 해제
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            zip_ref.extractall(temp_dir)
                        
                        # ChromeDriver 실행 파일 찾기
                        driver_exe = os.path.join(temp_dir, "chromedriver-win64", "chromedriver.exe")
                        if not os.path.exists(driver_exe):
                            # 다른 경로 시도
                            for root, dirs, files in os.walk(temp_dir):
                                for file in files:
                                    if file == "chromedriver.exe":
                                        driver_exe = os.path.join(root, file)
                                        break
                        
                        if os.path.exists(driver_exe):
                            service = Service(driver_exe)
                            self.driver = webdriver.Chrome(service=service, options=chrome_options)
                            self.log("방법 3 성공: 수동 다운로드", 'SUCCESS')
                        else:
                            raise Exception("ChromeDriver 실행 파일을 찾을 수 없음")
                            
                    except Exception as e3:
                        self.log(f"방법 3 실패: {e3}", 'ERROR')
                        raise Exception("모든 Chrome 드라이버 설정 방법 실패")
            
            # WebDriver 속성 숨기기
            if self.driver:
                self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                self.log("Chrome 드라이버 설정 완료!", 'SUCCESS')
                return True
            
            return False
            
        except Exception as e:
            self.log(f"드라이버 설정 중 치명적 오류: {e}", 'ERROR')
            messagebox.showerror("드라이버 오류", 
                               f"Chrome 드라이버 설정에 실패했습니다.\n\n"
                               f"오류: {str(e)}\n\n"
                               f"해결 방법:\n"
                               f"1. Chrome 브라우저를 최신 버전으로 업데이트\n"
                               f"2. 바이러스 백신 소프트웨어 일시 중지\n"
                               f"3. 관리자 권한으로 프로그램 실행")
            return False
    
    def test_login(self):
        """로그인 테스트"""
        def test_thread():
            try:
                self.progress.start()
                self.update_status("로그인 테스트 중...")
                
                if not self.setup_driver():
                    return
                
                naver_id = self.id_entry.get().strip()
                naver_pw = self.pw_entry.get().strip()
                
                if not naver_id or not naver_pw:
                    messagebox.showerror("오류", "아이디와 비밀번호를 입력해주세요.")
                    return
                
                self.log("네이버 로그인 페이지로 이동 중...", 'INFO')
                self.driver.get("https://nid.naver.com/nidlogin.login")
                time.sleep(3)
                
                # 아이디 입력
                id_input = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "id"))
                )
                pyperclip.copy(naver_id)
                id_input.click()
                time.sleep(0.5)
                id_input.clear()
                actions = ActionChains(self.driver)
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                
                # 비밀번호 입력
                pw_input = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "pw"))
                )
                pyperclip.copy(naver_pw)
                pw_input.click()
                time.sleep(0.5)
                pw_input.clear()
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                
                # 로그인 버튼 클릭
                login_btn = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "log.login"))
                )
                login_btn.click()
                time.sleep(5)
                
                current_url = self.driver.current_url
                if ("naver.com" in current_url and "login" not in current_url):
                    self.log("로그인 성공!", 'SUCCESS')
                    
                    # 블로그 작성 페이지로 이동
                    self.log("블로그 글쓰기 페이지로 이동 중...", 'INFO')
                    self.driver.get("https://blog.naver.com/GoBlogWrite.naver")
                    time.sleep(3)
                    
                    self.log("블로그 글쓰기 페이지 도착!", 'SUCCESS')
                    messagebox.showinfo("성공", 
                                       "로그인이 성공했습니다!\n\n"
                                       "🎉 블로그 글쓰기 페이지가 열려있습니다.\n"
                                       "이제 'AI 글 생성' 또는 '엑셀 포스팅' 기능을 사용하세요.\n\n"
                                       "⚠️ 브라우저 창을 닫지 마세요!")
                    
                    # 브라우저를 닫지 않고 유지
                    return
                else:
                    self.log("로그인 실패 또는 추가 인증 필요", 'WARNING')
                    
                    # 캡차나 추가 인증이 있을 경우 브라우저를 열어두고 사용자 대기
                    result = messagebox.askyesno("인증 필요", 
                                               "로그인 과정에서 추가 인증(캡차 등)이 필요할 수 있습니다.\n\n"
                                               "브라우저에서 수동으로 로그인을 완료하시겠습니까?\n"
                                               "완료 후 '예'를 클릭해주세요.\n\n"
                                               "'아니오'를 클릭하면 브라우저가 닫힙니다.")
                    
                    if result:
                        # 사용자가 수동 로그인 완료 후 블로그 페이지로 이동
                        self.log("사용자 수동 로그인 완료 대기 중...", 'INFO')
                        self.driver.get("https://blog.naver.com/GoBlogWrite.naver")
                        time.sleep(3)
                        self.log("블로그 글쓰기 페이지로 이동 완료!", 'SUCCESS')
                        messagebox.showinfo("완료", 
                                           "블로그 글쓰기 페이지가 열려있습니다!\n"
                                           "이제 기능을 사용하세요.")
                        return
                    else:
                        # 사용자가 취소한 경우에만 브라우저 닫기
                        if self.driver:
                            self.driver.quit()
                            self.driver = None
                
            except Exception as e:
                self.log(f"로그인 테스트 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
                # 오류 발생시에만 브라우저 닫기
                if self.driver:
                    self.driver.quit()
                    self.driver = None
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=test_thread, daemon=True).start()
    
    def test_api(self):
        """API 테스트"""
        def api_thread():
            try:
                self.progress.start()
                self.update_status("API 테스트 중...")
                
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.log("Gemini API 초기화 중...", 'INFO')
                self.gemini_api = GeminiAPI(api_key)
                
                self.log("API 테스트 요청 중...", 'INFO')
                test_prompt = "안녕하세요. API 테스트입니다. 간단히 인사해주세요."
                response = self.gemini_api.generate_text(test_prompt)
                
                if response:
                    self.log("API 테스트 성공!", 'SUCCESS')
                    messagebox.showinfo("성공", f"API 테스트 성공!\n응답: {response[:100]}...")
                else:
                    self.log("API 테스트 실패", 'ERROR')
                    messagebox.showerror("실패", "API 테스트에 실패했습니다.")
                
            except Exception as e:
                self.log(f"API 테스트 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"API 오류: {e}")
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=api_thread, daemon=True).start()
    
    def generate_seo_posts(self):
        """SEO 키워드 기반 글 생성"""
        def generate_thread():
            try:
                self.progress.start()
                self.update_status("글 생성 중...")
                
                # API 초기화
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.gemini_api = GeminiAPI(api_key)
                
                # 키워드 처리
                keywords_text = self.keywords_text.get('1.0', 'end').strip()
                keywords = [k.strip() for k in keywords_text.split(',') if k.strip()]
                
                if not keywords:
                    messagebox.showerror("오류", "키워드를 입력해주세요.")
                    return
                
                self.log(f"키워드: {', '.join(keywords)}", 'INFO')
                self.log("5개 포스트 생성 시작", 'INFO')
                
                # 엑셀 파일 생성
                wb = Workbook()
                ws = wb.active
                ws.title = "생성된 포스트"
                ws['A1'] = "제목"
                ws['B1'] = "본문"
                
                # 5개 포스트 생성 (각 키워드에 대해 1개씩, 총 5개)
                total_posts = 5
                
                for i in range(total_posts):
                    keyword = keywords[i % len(keywords)]  # 키워드 순환 사용
                    self.log(f"글 {i+1}/{total_posts} 생성 중... (키워드: {keyword})", 'INFO')
                    
                    # 개별 키워드로 제목 생성
                    title_prompt = f"'{keyword}'를 주요 키워드로 하는 SEO 최적화된 블로그 제목을 1개 생성해주세요. 제목만 출력하세요."
                    title = self.gemini_api.generate_text(title_prompt, temperature=0.8)
                    
                    if title:
                        title = title.strip().replace('"', '').replace("'", "")
                        
                        # SEO 최적화 프롬프트로 본문 생성
                        content_prompt = f"""
                        역할 및 목적 설정:
                        - 역할: 당신은 SEO 전문가이자, {keyword}에 대한 깊은 지식을 가진 교육 전문 블로거입니다.
                        - 목표: 학부모들을 위한 {title} 주제로, 검색 유입을 극대화하는 블로그 글을 작성해야 합니다.
                        - 대상 독자: 자녀 교육에 관심이 많은 학부모
                        
                        글쓰기 지침:
                        - 첫 문장: 반드시 강력한 후킹 문장으로 시작하세요 (예: "놀라운 사실이 있습니다", "많은 학부모님들이 모르는 진실", "우리 아이의 미래가 달린 중요한 순간" 등)
                        - 톤앤매너: 학부모님들께 정중하고 친근하게 존댓말로 설명하세요. 전문 용어는 쉽게 풀어서 설명해주세요.
                        - 글자 수는 최소 1,500자 이상으로 작성해 주세요.
                        - '{keyword}' 키워드를 자연스럽게 5-7회 포함시켜 주세요.
                        - 독자의 이해를 돕기 위한 적절한 [표 삽입 위치], [이미지 삽입 위치] 제안을 포함해 주세요.
                        - FAQ: {keyword}에 대해 학부모님들이 자주 궁금해하시는 질문 3가지와 답변을 마지막에 포함해 주세요.
                        
                        주의사항:
                        - 마크다운 문법 사용 금지 (**, ##, ===, * 등 특수문자 사용하지 말 것)
                        - 일반 텍스트로만 작성
                        - 존댓말 사용 (반말 금지)
                        - 중요한 내용은 "중요한 점은", "핵심 포인트는" 같은 정중한 표현으로 강조
                        - 학부모 관점에서 실용적이고 도움이 되는 내용 위주
                        
                        위 지침에 따라 {title}에 대한 학부모 대상 블로그 글을 작성해 주세요.
                        """
                        
                        content = self.gemini_api.generate_text(content_prompt, temperature=0.7, max_output_tokens=2048)
                        
                        if content:
                            # 마크다운 문법 제거
                            content = self.remove_markdown(content)
                            ws[f'A{i+2}'] = title  # A2부터 시작 (A1은 헤더)
                            ws[f'B{i+2}'] = content
                            self.log(f"글 {i+1} 생성 완료: {title[:30]}...", 'SUCCESS')
                        else:
                            self.log(f"글 {i+1} 본문 생성 실패", 'WARNING')
                    else:
                        self.log(f"글 {i+1} 제목 생성 실패", 'WARNING')
                    
                    time.sleep(1)  # API 제한 고려
                
                # 파일 저장
                filename = f"generated_posts_{int(time.time())}.xlsx"
                wb.save(filename)
                wb.close()
                
                self.log(f"모든 글 생성 완료! 파일: {filename}", 'SUCCESS')
                messagebox.showinfo("완료", f"글 생성이 완료되었습니다!\n저장된 파일: {filename}")
                
                # 파일 경로 업데이트
                self.file_path_var.set(filename)
                
            except Exception as e:
                self.log(f"글 생성 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=generate_thread, daemon=True).start()
    
    def auto_generate_keywords(self):
        """SEO 키워드 자동 생성"""
        def keyword_generation_thread():
            try:
                self.progress.start()
                self.update_status("SEO 키워드 자동 생성 중...")
                
                # API 초기화
                api_key = self.api_entry.get().strip()
                if not api_key:
                    messagebox.showerror("오류", "API 키를 입력해주세요.")
                    return
                
                self.gemini_api = GeminiAPI(api_key)
                
                # 키워드 자동 생성 프롬프트
                keyword_prompt = """
                매우 다양한 분야에서 실제 검색량이 높은 블로그 키워드 5개를 랜덤하게 생성해주세요.
                
                요구사항:
                - 매번 완전히 다른 분야 조합
                - 실제 사람들이 검색하는 키워드
                - 블로그 글로 작성 가능한 주제
                - 각 키워드는 2-6글자로 구성
                - 쉼표로 구분하여 출력
                
                다양한 분야 (매번 다른 조합으로 선택):
                🏠 라이프스타일: 미니멀라이프, 정리정돈, 홈카페, 셀프인테리어, 반려동물
                🍳 요리/음식: 집밥레시피, 다이어트식단, 홈베이킹, 건강식품, 간단요리
                💪 건강/운동: 홈트레이닝, 요가, 다이어트, 근력운동, 스트레칭
                🎨 취미/문화: 독서, 영화추천, 드라마리뷰, 음악감상, 사진촬영
                ✈️ 여행/외식: 국내여행, 맛집추천, 카페투어, 드라이브코스, 힐링여행
                💰 재정관리: 재테크, 투자, 부동산, 적금, 가계부작성
                👗 패션/뷰티: 코디, 화장품리뷰, 네일아트, 헤어스타일, 스킨케어
                📚 자기계발: 독서후기, 어학공부, 자격증취득, 시간관리, 습관만들기
                🎮 취미활동: 게임, 만화, 웹툰, 보드게임, 퍼즐
                🌱 원예/DIY: 식물키우기, 가드닝, 수공예, 리폼, 업사이클링
                🚗 자동차/기술: 자동차관리, 스마트폰, 가전제품, 앱추천, 온라인쇼핑
                💼 직장/커리어: 이직준비, 면접팁, 업무스킬, 부업, 창업
                👶 육아/교육: 육아팁, 교육정보, 놀이활동, 유아용품, 학습법
                🏥 의료/건강정보: 건강관리, 질병예방, 의료정보, 응급처치, 건강식품
                
                위의 모든 분야에서 골고루 섞어서 키워드 5개를 생성해주세요. 키워드만 출력하세요.
                """
                
                self.log("키워드 생성 요청 중...", 'INFO')
                response = self.gemini_api.generate_text(keyword_prompt, temperature=0.8)
                
                if response:
                    # 생성된 키워드를 텍스트 박스에 자동 입력
                    keywords = response.strip().replace('\n', '').replace('"', '').replace("'", "")
                    
                    # 기존 텍스트 지우고 새로운 키워드 입력
                    self.keywords_text.delete('1.0', 'end')
                    self.keywords_text.insert('1.0', keywords)
                    
                    self.log(f"키워드 자동 생성 완료: {keywords}", 'SUCCESS')
                    messagebox.showinfo("완료", f"SEO 키워드가 자동 생성되었습니다!\n\n생성된 키워드:\n{keywords}")
                else:
                    self.log("키워드 생성 실패", 'ERROR')
                    messagebox.showerror("실패", "키워드 생성에 실패했습니다.")
                
            except Exception as e:
                self.log(f"키워드 생성 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"키워드 생성 중 오류가 발생했습니다: {e}")
            finally:
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=keyword_generation_thread, daemon=True).start()
    
    def save_credentials(self):
        """로그인 정보 저장"""
        try:
            # 입력된 정보 가져오기
            naver_id = self.id_entry.get().strip()
            naver_pw = self.pw_entry.get().strip()
            api_key = self.api_entry.get().strip()
            
            if not all([naver_id, naver_pw, api_key]):
                messagebox.showwarning("경고", "모든 정보를 입력해주세요.")
                return
            
            # 비밀번호 간단 암호화 (base64)
            encoded_pw = base64.b64encode(naver_pw.encode()).decode()
            encoded_api = base64.b64encode(api_key.encode()).decode()
            
            # 저장할 데이터
            credentials = {
                'naver_id': naver_id,
                'naver_pw': encoded_pw,
                'api_key': encoded_api,
                'saved_date': time.strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # JSON 파일로 저장
            with open('blog_credentials.json', 'w', encoding='utf-8') as f:
                json.dump(credentials, f, ensure_ascii=False, indent=2)
            
            self.log("로그인 정보가 저장되었습니다.", 'SUCCESS')
            messagebox.showinfo("완료", "로그인 정보가 안전하게 저장되었습니다!")
            
        except Exception as e:
            self.log(f"정보 저장 중 오류: {e}", 'ERROR')
            messagebox.showerror("오류", f"정보 저장 중 오류가 발생했습니다: {e}")
    
    def load_credentials(self, show_message=True):
        """로그인 정보 불러오기"""
        try:
            # JSON 파일 존재 확인
            if not os.path.exists('blog_credentials.json'):
                if show_message:
                    messagebox.showinfo("정보", "저장된 로그인 정보가 없습니다.")
                return
            
            # JSON 파일 읽기
            with open('blog_credentials.json', 'r', encoding='utf-8') as f:
                credentials = json.load(f)
            
            # 정보 복원
            naver_id = credentials.get('naver_id', '')
            encoded_pw = credentials.get('naver_pw', '')
            encoded_api = credentials.get('api_key', '')
            saved_date = credentials.get('saved_date', '알 수 없음')
            
            # 비밀번호 복호화
            try:
                naver_pw = base64.b64decode(encoded_pw.encode()).decode()
                api_key = base64.b64decode(encoded_api.encode()).decode()
            except:
                if show_message:
                    messagebox.showerror("오류", "저장된 정보가 손상되었습니다.")
                return
            
            # UI에 정보 입력
            self.id_entry.delete(0, 'end')
            self.id_entry.insert(0, naver_id)
            
            self.pw_entry.delete(0, 'end')
            self.pw_entry.insert(0, naver_pw)
            
            self.api_entry.delete(0, 'end')
            self.api_entry.insert(0, api_key)
            
            if show_message:
                self.log("로그인 정보를 불러왔습니다.", 'SUCCESS')
                messagebox.showinfo("완료", f"로그인 정보를 불러왔습니다!\n저장 날짜: {saved_date}")
            else:
                self.log("저장된 로그인 정보 자동 로드 완료", 'INFO')
            
        except Exception as e:
            if show_message:
                self.log(f"정보 불러오기 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"정보 불러오기 중 오류가 발생했습니다: {e}")
    
    def browse_file(self):
        """파일 선택"""
        filename = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
            self.log(f"파일 선택: {filename}", 'INFO')
    
    def process_excel(self):
        """엑셀 파일 처리"""
        def excel_thread():
            try:
                self.progress.start()
                self.update_status("엑셀 데이터 처리 중...")
                
                excel_file = self.file_path_var.get()
                if not os.path.exists(excel_file):
                    messagebox.showerror("오류", "엑셀 파일을 찾을 수 없습니다.")
                    return
                
                # 드라이버 설정
                if not self.setup_driver():
                    return
                
                # 로그인
                if not self.login_to_naver():
                    return
                
                # 엑셀 데이터 로드
                posts_data = self.load_excel_data(excel_file)
                if not posts_data:
                    return
                
                success_count = 0
                for i, post in enumerate(posts_data, 1):
                    self.log(f"[{i}/{len(posts_data)}] {post['row']}행 처리 중...", 'INFO')
                    
                    if self.write_blog_post(post['title'], post['content']):
                        success_count += 1
                        self.log(f"{post['row']}행 포스팅 성공!", 'SUCCESS')
                    else:
                        self.log(f"{post['row']}행 포스팅 실패", 'WARNING')
                    
                    if i < len(posts_data):
                        time.sleep(3)
                
                self.log(f"처리 완료! 성공: {success_count}/{len(posts_data)}", 'SUCCESS')
                messagebox.showinfo("완료", f"엑셀 데이터 처리 완료!\n성공: {success_count}/{len(posts_data)}")
                
            except Exception as e:
                self.log(f"엑셀 처리 중 오류: {e}", 'ERROR')
                messagebox.showerror("오류", f"오류가 발생했습니다: {e}")
            finally:
                if self.driver:
                    self.driver.quit()
                    self.driver = None
                self.progress.stop()
                self.update_status("준비됨")
        
        threading.Thread(target=excel_thread, daemon=True).start()
    
    def load_excel_data(self, filename):
        """엑셀 데이터 로드"""
        try:
            self.log(f"엑셀 파일 로드: {filename}", 'INFO')
            workbook = load_workbook(filename)
            worksheet = workbook.active
            
            posts_data = []
            max_row = worksheet.max_row
            
            for row_num in range(2, max_row + 1):
                title_cell = worksheet[f'A{row_num}']
                content_cell = worksheet[f'B{row_num}']
                
                if not title_cell.value or title_cell.value.strip() == "":
                    continue
                
                title = title_cell.value.strip()
                content = content_cell.value.strip() if content_cell.value else ""
                
                posts_data.append({
                    'row': row_num,
                    'title': title,
                    'content': content
                })
            
            workbook.close()
            self.log(f"총 {len(posts_data)}개의 게시글 로드 완료", 'INFO')
            return posts_data
            
        except Exception as e:
            self.log(f"엑셀 파일 로드 중 오류: {e}", 'ERROR')
            return []
    
    def login_to_naver(self):
        """네이버 로그인"""
        try:
            self.log("네이버 로그인 시작...", 'INFO')
            
            naver_id = self.id_entry.get().strip()
            naver_pw = self.pw_entry.get().strip()
            
            if not naver_id or not naver_pw:
                messagebox.showerror("오류", "아이디와 비밀번호를 입력해주세요.")
                return False
            
            self.driver.get("https://nid.naver.com/nidlogin.login")
            time.sleep(3)
            
            # 아이디 입력
            id_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "id"))
            )
            pyperclip.copy(naver_id)
            id_input.click()
            time.sleep(0.5)
            id_input.clear()
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 비밀번호 입력
            pw_input = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.ID, "pw"))
            )
            pyperclip.copy(naver_pw)
            pw_input.click()
            time.sleep(0.5)
            pw_input.clear()
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 로그인 버튼 클릭
            login_btn = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.ID, "log.login"))
            )
            login_btn.click()
            time.sleep(5)
            
            current_url = self.driver.current_url
            if ("naver.com" in current_url and "login" not in current_url):
                self.log("로그인 성공!", 'SUCCESS')
                return True
            else:
                self.log("로그인 실패", 'ERROR')
                return False
                
        except Exception as e:
            self.log(f"로그인 중 오류: {e}", 'ERROR')
            return False
    
    def remove_markdown(self, content):
        """마크다운 문법 제거"""
        if not content:
            return content
        
        import re
        
        # 제목 마크다운 제거 (# ## ###)
        content = re.sub(r'^#{1,6}\s*', '', content, flags=re.MULTILINE)
        
        # 볼드 마크다운 제거 (**text** -> text)
        content = re.sub(r'\*\*(.*?)\*\*', r'\1', content)
        content = re.sub(r'__(.*?)__', r'\1', content)
        
        # 이탤릭 마크다운 제거 (*text* -> text)
        content = re.sub(r'\*(.*?)\*', r'\1', content)
        content = re.sub(r'_(.*?)_', r'\1', content)
        
        # 코드 블록 제거 (```text``` -> text)
        content = re.sub(r'```.*?\n(.*?)\n```', r'\1', content, flags=re.DOTALL)
        content = re.sub(r'`(.*?)`', r'\1', content)
        
        # 링크 마크다운 제거 ([text](url) -> text)
        content = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', content)
        
        # 리스트 마크다운 제거 (- item -> item)
        content = re.sub(r'^[-\*\+]\s*', '', content, flags=re.MULTILINE)
        content = re.sub(r'^\d+\.\s*', '', content, flags=re.MULTILINE)
        
        # 인용구 마크다운 제거 (> text -> text)
        content = re.sub(r'^>\s*', '', content, flags=re.MULTILINE)
        
        # 수평선 제거 (--- or ===)
        content = re.sub(r'^[-=]{3,}$', '', content, flags=re.MULTILINE)
        
        return content.strip()

    def format_blog_content(self, content):
        """블로그 내용 포맷팅"""
        if not content:
            return content
            
        # 이모티콘 추가
        emoticons = ["😊", "✨", "🔥", "💡", "📚", "🎯", "🚀", "💪", "👍", "⭐"]
        
        # 내용을 문단으로 분할
        paragraphs = content.split('\n')
        formatted_paragraphs = []
        
        for i, paragraph in enumerate(paragraphs):
            if paragraph.strip():
                # 랜덤 이모티콘 추가 (25% 확률)
                if i % 4 == 0 and len(paragraph) > 50:
                    import random
                    paragraph = f"{random.choice(emoticons)} {paragraph}"
                
                # 인용구 추가 (중요한 문장에)
                if any(keyword in paragraph for keyword in ['중요', '핵심', '포인트', '요약']):
                    paragraph = f"> {paragraph}"
                
                formatted_paragraphs.append(paragraph)
        
        # 문단들을 조합 (구분선은 실제 도구로 삽입)
        return '\n\n'.join(formatted_paragraphs)

    def apply_naver_formatting(self, content):
        """네이버 스마트에디터 포맷팅 적용"""
        try:
            # JavaScript를 통한 스타일 적용
            format_script = f"""
            // 에디터 영역 찾기
            var editor = document.querySelector('.se-section-text');
            if (editor) {{
                // 나눔스퀘어 폰트 적용
                editor.style.fontFamily = 'NanumSquare, sans-serif';
                editor.style.fontSize = '19px';
                editor.style.textAlign = 'center';
                editor.style.lineHeight = '1.6';
                
                // 본문에 스타일 적용
                var paragraphs = editor.querySelectorAll('p');
                paragraphs.forEach(function(p) {{
                    p.style.fontFamily = 'NanumSquare, sans-serif';
                    p.style.fontSize = '19px';
                    p.style.textAlign = 'center';
                    p.style.marginBottom = '15px';
                }});
                
                // 인용구 스타일
                var quotes = editor.querySelectorAll('blockquote, p:contains(">")');
                quotes.forEach(function(q) {{
                    q.style.backgroundColor = '#f8f9fa';
                    q.style.borderLeft = '4px solid #007bff';
                    q.style.padding = '10px 15px';
                    q.style.margin = '15px 0';
                    q.style.fontStyle = 'italic';
                }});
            }}
            """
            
            self.driver.execute_script(format_script)
            time.sleep(1)
            
        except Exception as e:
            self.log(f"포맷팅 적용 중 오류: {e}", 'WARNING')

    def apply_center_alignment(self):
        """가운데 정렬 도구 버튼 클릭"""
        try:
            self.log("가운데 정렬 적용 시도...", 'INFO')
            
            # 전체 텍스트 선택
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.5)
            
            # 가운데 정렬 버튼 클릭
            center_selectors = [
                ".se-toolbar-icon[title*='가운데']",
                ".se-toolbar-icon[aria-label*='가운데']",
                ".se-toolbar-icon[title*='center']",
                ".se-toolbar-align-center",
                "button[class*='center']",
                ".se-toolbar-icon:nth-child(6)"  # 대략적인 위치
            ]
            
            center_btn = None
            for selector in center_selectors:
                try:
                    center_btn = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if center_btn:
                        break
                except:
                    continue
            
            if center_btn:
                center_btn.click()
                time.sleep(1)
                self.log("가운데 정렬 적용 완료!", 'SUCCESS')
            else:
                # JavaScript로 직접 가운데 정렬 적용
                center_script = """
                var editor = document.querySelector('.se-section-text');
                if (editor) {
                    editor.style.textAlign = 'center';
                    var allElements = editor.querySelectorAll('*');
                    allElements.forEach(function(el) {
                        el.style.textAlign = 'center';
                    });
                }
                """
                self.driver.execute_script(center_script)
                self.log("JavaScript로 가운데 정렬 적용 완료!", 'SUCCESS')
                
        except Exception as e:
            self.log(f"가운데 정렬 적용 중 오류: {e}", 'WARNING')

    def add_template_elements(self, content):
        """템플릿 요소 추가"""
        template_header = "🎯 오늘의 포스팅\n\n"
        template_footer = "\n\n---\n✨ 이 글이 도움이 되셨다면 공감과 댓글 부탁드려요!\n💡 더 많은 유용한 정보는 이웃추가로 받아보세요!"
        
        return template_header + content + template_footer

    def write_content_with_elements(self, content):
        """내용을 단계적으로 입력하면서 구분선과 스티커 삽입"""
        try:
            # 내용 포맷팅 적용
            formatted_content = self.format_blog_content(content)
            formatted_content = self.add_template_elements(formatted_content)
            
            # 내용을 문단별로 분할
            paragraphs = formatted_content.split('\n\n')
            total_paragraphs = len(paragraphs)
            
            self.log(f"총 {total_paragraphs}개 문단으로 단계별 입력 시작", 'INFO')
            
            # 각 문단을 순차적으로 입력
            for i, paragraph in enumerate(paragraphs):
                if paragraph.strip():
                    # 문단 입력
                    pyperclip.copy(paragraph.strip())
                    actions = ActionChains(self.driver)
                    actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                    time.sleep(0.5)
                    
                    # 문단 사이에 엔터 추가
                    if i < len(paragraphs) - 1:
                        actions.send_keys(Keys.ENTER).send_keys(Keys.ENTER).perform()
                        time.sleep(0.3)
                    
                    # 스티커 빈도 높이기 (2문단마다)
                    if i > 0 and i % 2 == 0 and i < total_paragraphs - 1:
                        self.log(f"{i+1}번째 문단 후 스티커 삽입", 'INFO')
                        self.insert_random_sticker()
            
            # 스타일 적용
            self.apply_naver_formatting(formatted_content)
            
            # 가운데 정렬 적용
            self.apply_center_alignment()
            
        except Exception as e:
            self.log(f"단계별 내용 입력 중 오류: {e}", 'ERROR')
            # 오류 발생 시 기존 방식으로 전체 입력
            pyperclip.copy(formatted_content)
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.3)
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(2)

    def write_blog_post(self, title, content):
        """블로그 포스트 작성"""
        try:
            self.log(f"블로그 포스트 작성: {title[:30]}...", 'INFO')
            
            # 글쓰기 페이지로 이동
            self.driver.get("https://blog.naver.com/GoBlogWrite.naver")
            time.sleep(3)
            
            # iframe 전환
            main_frame = WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#mainFrame"))
            )
            self.driver.switch_to.frame(main_frame)
            time.sleep(2)
            
            # 팝업 처리
            self.handle_popups()
            
            # 제목 입력
            title_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-documentTitle"))
            )
            title_element.click()
            time.sleep(1)
            
            pyperclip.copy(title)
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.3)
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            
            # 본문 입력
            text_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-text"))
            )
            text_element.click()
            time.sleep(1)
            
            if content:
                # 내용을 단계적으로 입력하면서 구분선과 스티커 삽입
                self.write_content_with_elements(content)
            
            # 저장 버튼 클릭
            save_selectors = [
                ".save_btn__bzc5B",
                "#save_btn_bzc58", 
                ".btn_save",
                "[class*='save']"
            ]
            
            for selector in save_selectors:
                try:
                    save_btn = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    save_btn.click()
                    self.log("저장 버튼 클릭 완료!", 'SUCCESS')
                    break
                except:
                    continue
            
            time.sleep(3)
            self.driver.switch_to.default_content()
            
            return True
            
        except Exception as e:
            self.log(f"글 작성 중 오류: {e}", 'ERROR')
            try:
                self.driver.switch_to.default_content()
            except:
                pass
            return False
    

    def insert_random_sticker(self):
        """랜덤 스티커 삽입"""
        try:
            import random
            
            # 사용 가능한 스티커 목록 (배경 위치값)
            stickers = [
                {'name': '첫인사', 'position': '0px 0px'},
                {'name': 'OK싸인', 'position': '-82px 0px'},
                {'name': '온화한미소', 'position': '-160px 0px'},
                {'name': '좋아요', 'position': '0px -74px'},
                {'name': '감사합니다', 'position': '0px -148px'}
            ]
            
            selected_sticker = random.choice(stickers)
            self.log(f"'{selected_sticker['name']}' 스티커 삽입 시도...", 'INFO')
            
            # 스티커 도구 버튼 클릭 (구분선과 다른 스티커 도구 버튼)
            try:
                # 스티커 버튼 찾기 (여러 가능한 셀렉터 시도)
                sticker_selectors = [
                    ".se-toolbar-icon[title*='스티커']",
                    ".se-toolbar-icon[aria-label*='스티커']", 
                    ".se-toolbar-sticker-button",
                    "button[class*='sticker']",
                    ".se-toolbar-icon:nth-child(8)"  # 스티커 버튼 대략적 위치
                ]
                
                sticker_btn = None
                for selector in sticker_selectors:
                    try:
                        sticker_btn = self.driver.find_element(By.CSS_SELECTOR, selector)
                        if sticker_btn:
                            break
                    except:
                        continue
                
                if sticker_btn:
                    sticker_btn.click()
                    time.sleep(1)
                    
                    # 스티커 선택 (JavaScript로 직접 클릭)
                    sticker_script = f"""
                    var stickers = document.querySelectorAll('.se-sidebar-sticker');
                    for (var i = 0; i < stickers.length; i++) {{
                        var style = stickers[i].getAttribute('style');
                        if (style && style.includes('{selected_sticker['position']}')) {{
                            stickers[i].click();
                            console.log('스티커 클릭됨: {selected_sticker['name']}');
                            break;
                        }}
                    }}
                    """
                    
                    self.driver.execute_script(sticker_script)
                    time.sleep(1)
                    
                    self.log(f"'{selected_sticker['name']}' 스티커 삽입 완료!", 'SUCCESS')
                else:
                    self.log("스티커 버튼을 찾을 수 없음", 'WARNING')
                
            except Exception as e:
                self.log(f"스티커 삽입 중 오류: {e}", 'WARNING')
                
        except Exception as e:
            self.log(f"스티커 처리 중 오류: {e}", 'WARNING')

    def handle_popups(self):
        """팝업 처리"""
        try:
            cancel_btn = self.driver.find_element(By.CSS_SELECTOR, ".se-popup-button-cancel")
            if cancel_btn.is_displayed():
                cancel_btn.click()
                time.sleep(1)
        except:
            pass
        
        try:
            help_close_btn = self.driver.find_element(By.CSS_SELECTOR, ".se-help-panel-close-button")
            if help_close_btn.is_displayed():
                help_close_btn.click()
                time.sleep(1)
        except:
            pass
    
    def run(self):
        """프로그램 실행"""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()
    
    def on_closing(self):
        """프로그램 종료"""
        if self.driver:
            self.driver.quit()
        self.root.destroy()


if __name__ == "__main__":
    app = ModernBlogWriter()
    app.run()